import pandas as pd
import openpyxl
import tempfile
import os

# Simple test
data = {'A': [1, 2], 'B': [3, 4]}
df = pd.DataFrame(data)

with tempfile.TemporaryDirectory() as temp_dir:
    temp_xlsx = os.path.join(temp_dir, 'temp.xlsx')
    final_xlsm = os.path.join(temp_dir, 'final.xlsm')
    
    # Create XLSX
    df.to_excel(temp_xlsx, index=False)
    
    # Load and save as XLSM
    wb = openpyxl.load_workbook(temp_xlsx)
    wb.code_name = "ThisWorkbook"
    wb.save(final_xlsm)
    
    # Verify
    wb2 = openpyxl.load_workbook(final_xlsm)
    print(f"Success: XLSM file created and readable")
    print(f"Rows: {wb2.active.max_row}, Cols: {wb2.active.max_column}")
    print(f"Code name: {wb2.code_name}")
