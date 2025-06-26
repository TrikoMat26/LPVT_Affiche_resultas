import os
import sys
import re
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from bs4 import BeautifulSoup
import pandas as pd
from typing import Dict, List, Any, Set, Optional, Tuple
import io
import subprocess
from Affiche_resultats import traiter_repertoire_serie, ProgressWindow 
import webbrowser
import openpyxl
from openpyxl.styles import Font 
import traceback
import configparser 
from pathlib import Path 

# Dictionnaire des bornes avec les noms de colonnes exacts
bornes_alims_seq01 = {
    "alim 24VDC +16V": (15.9, 16.3), "alim 24VDC -16V": (-16.3, -15.9),
    "alim 24VDC +5V": (5.2, 5.5), "alim 24VDC -5V": (-5.5, -5.2),
    "alim 115VAC +16V": (16, 16.3), "alim 115VAC -16V": (-16.36, -15),
    "1.9Un en 19VDC": (-16.6, -15), "1.9Un en 115VAC": (-16.6, -15),
}

CONFIG_FILE_NAME = "lpvt_analyzer_config.ini"

def get_config_path() -> Path:
    if sys.platform == "win32":
        app_data_dir = Path(os.getenv('APPDATA', Path.home() / "AppData" / "Roaming"))
    else:
        app_data_dir = Path.home() / ".config"
    app_config_dir = app_data_dir / "LPVTTestAnalyzer"
    app_config_dir.mkdir(parents=True, exist_ok=True)
    return app_config_dir / CONFIG_FILE_NAME

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def extraire_defauts_precision_transfert(html_content):
    resultats = []
    soup = BeautifulSoup(html_content, "html.parser")
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        i = 0
        while i < len(rows):
            tds = rows[i].find_all("td")
            if len(tds) == 2 and "Status:" in tds[0].get_text(strip=True):
                status = tds[1].get_text(strip=True)
                if status.lower() == "failed":
                    voie_text_element = None
                    precision = None
                    for j in range(i + 1, min(i + 5, len(rows))):
                        tds2 = rows[j].find_all("td")
                        if len(tds2) == 2:
                            label_text_inner = tds2[0].get_text(strip=True)
                            value_text_inner = tds2[1].get_text(strip=True)
                            if "ROUE CODEUSE" in label_text_inner and "GAMME" in label_text_inner and "Voie" in label_text_inner:
                                voie_text_element = label_text_inner
                            if "Valeur réelle injectée" in label_text_inner and "/" in value_text_inner:
                                parts = value_text_inner.split('/')
                                if len(parts) >= 4:
                                    precision_candidate = parts[-1].strip()
                                    try:
                                        float(precision_candidate.replace('%', '').strip())
                                        precision = precision_candidate
                                    except ValueError: pass
                    if voie_text_element and precision:
                        m = re.search(r'ROUE CODEUSE = ([^ \.]+) .. GAMME = ([^ \.]+) .. Voie (\w+)', voie_text_element)
                        if m:
                            rc, gamme, voie_char = m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
                            nom = f"RC:{rc}/G:{gamme}-Voie {voie_char}"
                            resultats.append((nom, precision))
            i += 1
    return resultats

class ModernStatsTestsWindow:
    def __init__(self):
        self.configurer_encodage()
        self.fichiers_seq01, self.fichiers_seq02 = [], []
        self.tests_disponibles, self.tests_selectionnes, self.data_tests = [], [], {}
        self.repertoire_parent = ""
        self.config_file_path = get_config_path()
        self.charger_configuration()

        self.root = tk.Tk()
        self.root.title("LPVT Test Analyzer - Interface Moderne")
        self.root.geometry("1200x1100")
        self.root.minsize(1000, 800) # Augmenté la hauteur minimale aussi
        
        self.setup_styles()
        self.creer_interface_moderne()
        self.creer_liste_tests_predefinies()

        if self.repertoire_parent and os.path.isdir(self.repertoire_parent):
            self.mettre_a_jour_affichage_repertoire(self.repertoire_parent)
            self.charger_fichiers_seq(self.repertoire_parent)
        else:
            self.repertoire_parent = ""
            self.mettre_a_jour_affichage_repertoire("")

    def charger_configuration(self):
        config = configparser.ConfigParser()
        if self.config_file_path.exists():
            try:
                config.read(self.config_file_path)
                if 'General' in config and 'last_directory' in config['General']:
                    last_dir = config['General']['last_directory']
                    if os.path.isdir(last_dir):
                        self.repertoire_parent = last_dir
                    else: self.repertoire_parent = ""
                else: self.repertoire_parent = ""
            except Exception: self.repertoire_parent = ""
        else: self.repertoire_parent = ""

    def sauvegarder_configuration(self):
        if not self.repertoire_parent: return
        config = configparser.ConfigParser()
        config['General'] = {'last_directory': self.repertoire_parent}
        try:
            with open(self.config_file_path, 'w') as configfile: config.write(configfile)
        except Exception as e: print(f"Erreur sauvegarde config: {e}")

    def mettre_a_jour_affichage_repertoire(self, repertoire_path: str):
        short_path = ("..." + repertoire_path[-57:]) if len(repertoire_path) > 60 else repertoire_path
        if not repertoire_path: short_path = "Aucun répertoire sélectionné"
        self.dir_path_label.config(text=short_path, 
                                   foreground=self.colors['text_primary'] if repertoire_path else self.colors['text_muted'])

    def setup_styles(self): # Identique à votre version précédente
        style = ttk.Style()
        style.theme_use('clam')
        self.colors = {
            'bg_primary': '#fafafa', 'bg_secondary': '#ffffff', 'bg_accent': '#f5f7fa',
            'text_primary': '#2d3748', 'text_secondary': '#718096', 'text_muted': '#a0aec0',
            'border': '#e2e8f0', 'primary': '#4a5568', 'success': '#38a169',
            'warning': '#ed8936', 'error': '#e53e3e', 'info': '#3182ce', 'accent': '#667eea'
        }
        style.configure('TFrame', background=self.colors['bg_primary'])
        style.configure('Card.TFrame', background=self.colors['bg_secondary'], relief='flat', borderwidth=1)
        style.configure('TLabel', background=self.colors['bg_primary'], foreground=self.colors['text_primary'], font=('Segoe UI', 10))
        style.configure('Title.TLabel', font=('Segoe UI', 18, 'bold'), foreground=self.colors['text_primary'], background=self.colors['bg_primary'])
        style.configure('Subtitle.TLabel', font=('Segoe UI', 12, 'bold'), foreground=self.colors['text_secondary'], background=self.colors['bg_primary'])
        style.configure('Caption.TLabel', font=('Segoe UI', 9), foreground=self.colors['text_muted'], background=self.colors['bg_primary'])
        style.configure('Status.TLabel', font=('Segoe UI', 9), foreground=self.colors['info'], background=self.colors['bg_primary'])
        style.configure('Primary.TButton', font=('Segoe UI', 10, 'bold'), padding=(25, 12), relief='flat', borderwidth=0, background=self.colors['primary'], foreground='white', focuscolor='none')
        style.configure('Success.TButton', font=('Segoe UI', 10, 'bold'), padding=(25, 12), relief='flat', borderwidth=0, background=self.colors['success'], foreground='white', focuscolor='none')
        style.configure('Info.TButton', font=('Segoe UI', 10), padding=(15, 8), relief='flat', borderwidth=0, background=self.colors['info'], foreground='white', focuscolor='none')
        style.configure('Secondary.TButton', font=('Segoe UI', 10), padding=(15, 8), relief='flat', borderwidth=1, background=self.colors['bg_secondary'], foreground=self.colors['text_primary'], focuscolor='none')
        style.map('Primary.TButton', background=[('active', '#3a4653')])
        style.map('Success.TButton', background=[('active', '#2f7d32')])
        style.map('Info.TButton', background=[('active', '#2c5aa0')])
        style.map('Secondary.TButton', background=[('active', self.colors['bg_accent']), ('pressed', self.colors['border'])])
        style.configure('Modern.Treeview', font=('Segoe UI', 10), rowheight=30, background=self.colors['bg_secondary'], foreground=self.colors['text_primary'], fieldbackground=self.colors['bg_secondary'], borderwidth=0, relief='flat')
        style.configure('Modern.Treeview.Heading', font=('Segoe UI', 10, 'bold'), background=self.colors['bg_accent'], foreground=self.colors['text_primary'], relief='flat', borderwidth=0)
        style.map('Modern.Treeview', background=[('selected', self.colors['accent'])], foreground=[('selected', 'white')])
        style.configure('Modern.TEntry', fieldbackground=self.colors['bg_secondary'], borderwidth=1, relief='solid', bordercolor=self.colors['border'], font=('Segoe UI', 10), padding=(10, 8))
        style.map('Modern.TEntry', bordercolor=[('focus', self.colors['accent'])])

    def creer_interface_moderne(self): # Identique à votre version précédente
        self.root.configure(bg=self.colors['bg_primary'])
        main_container = tk.Frame(self.root, bg=self.colors['bg_primary'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        header_card = ttk.Frame(main_container, style='Card.TFrame'); header_card.pack(fill=tk.X, pady=(0,20))
        header_content = ttk.Frame(header_card); header_content.pack(fill=tk.X, padx=25, pady=20)
        title_section = ttk.Frame(header_content); title_section.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(title_section, text="LPVT Test Analyzer", style='Title.TLabel').pack(anchor='w')
        ttk.Label(title_section, text="Analyse statistique des tests SEQ-01 et SEQ-02", style='Subtitle.TLabel').pack(anchor='w', pady=(2,0))
        header_actions = ttk.Frame(header_content); header_actions.pack(side=tk.RIGHT)
        ttk.Button(header_actions, text="💡 Aide", command=self.show_help, style='Info.TButton').pack(side=tk.RIGHT, padx=(10,0))
        dir_card = ttk.Frame(main_container, style='Card.TFrame'); dir_card.pack(fill=tk.X, pady=(0,20))
        dir_content = ttk.Frame(dir_card); dir_content.pack(fill=tk.X, padx=25, pady=20)
        ttk.Label(dir_content, text="📁 Répertoire source", style='Subtitle.TLabel').pack(anchor='w', pady=(0,15))
        dir_row = ttk.Frame(dir_content); dir_row.pack(fill=tk.X)
        ttk.Button(dir_row, text="Parcourir...", command=self.selectionner_repertoire, style='Secondary.TButton').pack(side=tk.LEFT, padx=(0,15))
        path_container = ttk.Frame(dir_row); path_container.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.dir_path_label = ttk.Label(path_container, text="Aucun répertoire sélectionné", style='Caption.TLabel'); self.dir_path_label.pack(anchor='w')
        self.files_info_label = ttk.Label(dir_content, text="", style='Status.TLabel'); self.files_info_label.pack(anchor='w', pady=(10,0))
        tests_card = ttk.Frame(main_container, style='Card.TFrame'); tests_card.pack(fill=tk.BOTH, expand=True, pady=(0,20))
        tests_content = ttk.Frame(tests_card); tests_content.pack(fill=tk.BOTH, expand=True, padx=25, pady=20)
        tests_header = ttk.Frame(tests_content); tests_header.pack(fill=tk.X, pady=(0,15))
        ttk.Label(tests_header, text="⚡ Sélection des tests à analyser", style='Subtitle.TLabel').pack(side=tk.LEFT)
        options_frame = ttk.Frame(tests_header); options_frame.pack(side=tk.RIGHT)
        self.tri_chrono_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="📅 Tri chronologique par n° de série", variable=self.tri_chrono_var).pack()
        controls_row = ttk.Frame(tests_content); controls_row.pack(fill=tk.X, pady=(0,15))
        selection_buttons = ttk.Frame(controls_row); selection_buttons.pack(side=tk.LEFT)
        ttk.Button(selection_buttons, text="Tout sélectionner", command=self.selectionner_tous_tests, style='Secondary.TButton').pack(side=tk.LEFT, padx=(0,8))
        ttk.Button(selection_buttons, text="Tout désélectionner", command=self.deselectionner_tous_tests, style='Secondary.TButton').pack(side=tk.LEFT)
        search_container = ttk.Frame(controls_row); search_container.pack(side=tk.RIGHT)
        ttk.Label(search_container, text="🔍", font=('Segoe UI', 12)).pack(side=tk.LEFT, padx=(0,5))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_container, textvariable=self.search_var, style='Modern.TEntry', width=25); search_entry.pack(side=tk.LEFT)
        search_entry.bind('<KeyRelease>', self.filter_tests_list)
        list_container = ttk.Frame(tests_content); list_container.pack(fill=tk.BOTH, expand=True)
        self.tests_tree = ttk.Treeview(list_container, selectmode='extended', columns=('category',), show='tree headings', style='Modern.Treeview', height=15) # Hauteur augmentée
        self.tests_tree.heading('#0', text='Test', anchor='w'); self.tests_tree.heading('category', text='Catégorie', anchor='w')
        self.tests_tree.column('#0', width=300, minwidth=200, stretch=tk.YES); self.tests_tree.column('category', width=200, minwidth=150, stretch=tk.YES)
        scrollbar = ttk.Scrollbar(list_container, orient=tk.VERTICAL, command=self.tests_tree.yview); scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tests_tree.configure(yscrollcommand=scrollbar.set); self.tests_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        actions_card = ttk.Frame(main_container, style='Card.TFrame'); actions_card.pack(fill=tk.X)
        actions_content = ttk.Frame(actions_card); actions_content.pack(fill=tk.X, padx=25, pady=20)
        actions_row = ttk.Frame(actions_content); actions_row.pack(fill=tk.X)
        ttk.Button(actions_row, text="📊 Générer le rapport statistique", command=self.generer_statistiques, style='Success.TButton').pack(side=tk.LEFT, padx=(0,15))
        ttk.Button(actions_row, text="📄 Générer les rapports détaillés", command=self.generer_rapports_detailles, style='Primary.TButton').pack(side=tk.LEFT)
        status_container = ttk.Frame(actions_content); status_container.pack(fill=tk.X, pady=(15,0))
        status_separator = ttk.Frame(status_container, height=1, style='Card.TFrame'); status_separator.pack(fill=tk.X, pady=(0,10))
        self.status_label = ttk.Label(status_container, text="✅ Prêt à traiter vos fichiers", style='Status.TLabel'); self.status_label.pack(side=tk.LEFT)
        ttk.Label(status_container, text="LPVT Analyzer v2.2", style='Caption.TLabel').pack(side=tk.RIGHT)

    def show_help(self): # Identique à votre version précédente
        help_window = tk.Toplevel(self.root); help_window.title("Aide - LPVT Test Analyzer"); help_window.geometry("650x550")
        help_window.configure(bg=self.colors['bg_primary']); help_window.transient(self.root); help_window.grab_set()
        help_window.geometry("+%d+%d" % (self.root.winfo_rootx()+50, self.root.winfo_rooty()+50))
        main_frame = ttk.Frame(help_window, style='Card.TFrame'); main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        content_frame = ttk.Frame(main_frame); content_frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)
        ttk.Label(content_frame, text="💡 Guide d'utilisation", style='Title.TLabel').pack(anchor='w', pady=(0,20))
        help_content = """🎯 OBJECTIF\nAnalyser les rapports de test HTML SEQ-01 et SEQ-02 pour générer des statistiques et des rapports détaillés.\n\n📋 ÉTAPES D'UTILISATION\n\n1️⃣ SÉLECTION DU RÉPERTOIRE\n   • Cliquez sur "Parcourir..." et sélectionnez le dossier parent contenant les sous-dossiers des numéros de série.\n   • Structure attendue: Dossier_Parent/SN12345/fichiers_HTML\n\n2️⃣ SÉLECTION DES TESTS\n   • Choisissez les paramètres à analyser. Utilisez la recherche 🔍 pour filtrer.\n\n3️⃣ GÉNÉRATION DES RAPPORTS\n   📊 Rapport Statistique: Crée un fichier Excel (.xlsm). Double-cliquez sur un N° de Série (colonne A) pour ouvrir son dossier.\n   📄 Rapports Détaillés: Génère un rapport texte par N° de série.\n\n⚙️ OPTIONS\n   📅 Tri chronologique: Organise les résultats par N° de série et date/heure.\n\n💡 CONSEILS\n   • Valeurs hors spécifications en rouge dans Excel.\n   • Rapports ouverts automatiquement.\n   • Barre de progression pour les tâches longues."""
        text_frame = ttk.Frame(content_frame); text_frame.pack(fill=tk.BOTH, expand=True)
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=('Segoe UI', 10), bg=self.colors['bg_secondary'], fg=self.colors['text_primary'], relief='flat', borderwidth=0, padx=15, pady=15)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview); text_widget.configure(yscrollcommand=scrollbar.set)
        text_widget.insert('1.0', help_content.strip()); text_widget.configure(state='disabled')
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y); text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        button_frame = ttk.Frame(content_frame); button_frame.pack(fill=tk.X, pady=(15,0))
        ttk.Button(button_frame, text="Fermer", command=help_window.destroy, style='Primary.TButton').pack(side=tk.RIGHT)

    def update_status(self, message): # Identique
        self.status_label.config(text=message)
        self.root.update_idletasks()

    def filter_tests_list(self, event=None): # Identique
        search_term = self.search_var.get().lower()
        for item in self.tests_tree.get_children(): self.tests_tree.delete(item)
        categories = {'seq01_24vdc': '🔋 Alimentations 24VDC (SEQ-01)', 'seq01_115vac': '⚡ Alimentations 115VAC (SEQ-01)', 
                      'seq01_resistances': '🔧 Résistances (SEQ-01)', 'seq02_transfert': '📊 Tests de transfert (SEQ-02)', 
                      'seq02_precision_transfert': '🎯 Précision transfert (SEQ-02)'}
        filtered_by_category = {}
        for nom_complet, test_parent, identifiant in self.tests_disponibles:
            if search_term in nom_complet.lower() or search_term in test_parent.lower():
                if test_parent not in filtered_by_category: filtered_by_category[test_parent] = []
                filtered_by_category[test_parent].append((nom_complet, identifiant))
        for category_id, tests_in_cat in filtered_by_category.items():
            category_name = categories.get(category_id, category_id.replace("_", " ").title())
            category_node = self.tests_tree.insert('', 'end', text=category_name, values=(f'category_{category_id}',), open=True)
            for nom_complet, identifiant in tests_in_cat:
                self.tests_tree.insert(category_node, 'end', text=f"  {nom_complet}", values=(identifiant,))

    def selectionner_tous_tests(self): # Identique
        for category_node_id in self.tests_tree.get_children():
            for test_node_id in self.tests_tree.get_children(category_node_id):
                self.tests_tree.selection_add(test_node_id)

    def deselectionner_tous_tests(self): # Identique
        self.tests_tree.selection_remove(self.tests_tree.get_children())
        for category_node_id in self.tests_tree.get_children():
            for test_node_id in self.tests_tree.get_children(category_node_id):
                self.tests_tree.selection_remove(test_node_id)

    def configurer_encodage(self): # Identique
        try:
            if sys.stdout and hasattr(sys.stdout, 'encoding') and sys.stdout.encoding != 'utf-8':
                sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        except (AttributeError, TypeError): pass

    def creer_liste_tests_predefinies(self): # Identique
        self.tests_disponibles = [
            ("alim 24VDC +16V", "seq01_24vdc", "24VDC_+16V"), ("alim 24VDC -16V", "seq01_24vdc", "24VDC_-16V"),
            ("alim 24VDC +5V", "seq01_24vdc", "24VDC_+5V"), ("alim 24VDC -5V", "seq01_24vdc", "24VDC_-5V"),
            ("alim 115VAC +16V", "seq01_115vac", "115VAC_+16V"), ("alim 115VAC -16V", "seq01_115vac", "115VAC_-16V"),
            ("R46 calculée", "seq01_resistances", "R46_calculee"), ("R46 à monter", "seq01_resistances", "R46_monter"),
            ("R47 calculée", "seq01_resistances", "R47_calculee"), ("R47 à monter", "seq01_resistances", "R47_monter"),
            ("R48 calculée", "seq01_resistances", "R48_calculee"), ("R48 à monter", "seq01_resistances", "R48_monter"),
            ("1.9Un en 19VDC", "seq02_transfert", "Test_19VDC"), ("1.9Un en 115VAC", "seq02_transfert", "Test_115VAC"),
            ("Défauts précision transfert", "seq02_precision_transfert", "precision_transfert_defauts"),
        ]
        self.filter_tests_list()

    def selectionner_repertoire(self): # Modifiée pour sauvegarder
        initial_dir = self.repertoire_parent if self.repertoire_parent and os.path.isdir(self.repertoire_parent) else os.path.expanduser("~")
        repertoire = filedialog.askdirectory(title="Sélectionnez le répertoire source", initialdir=initial_dir)
        if not repertoire: return
        self.repertoire_parent = repertoire
        self.mettre_a_jour_affichage_repertoire(self.repertoire_parent)
        self.update_status(f"🔍 Analyse du répertoire {os.path.basename(repertoire)}...")
        self.charger_fichiers_seq(repertoire)
        self.sauvegarder_configuration()

    def charger_fichiers_seq(self, repertoire): # Identique
        self.update_status("🔍 Recherche des fichiers SEQ-01/SEQ-02...")
        self.fichiers_seq01, self.fichiers_seq02 = [], []
        for dossier_racine, _, fichiers in os.walk(repertoire):
            for fichier in fichiers:
                if fichier.startswith("SEQ-01") and fichier.lower().endswith(".html"): self.fichiers_seq01.append(os.path.join(dossier_racine, fichier))
                elif fichier.startswith("SEQ-02") and fichier.lower().endswith(".html"): self.fichiers_seq02.append(os.path.join(dossier_racine, fichier))
        total_fichiers = len(self.fichiers_seq01) + len(self.fichiers_seq02)
        if total_fichiers == 0:
            self.files_info_label.config(text="❌ Aucun fichier SEQ-01 ou SEQ-02 trouvé.", foreground=self.colors['error'])
            self.update_status("❌ Aucun fichier trouvé"); return
        self.files_info_label.config(text=f"✅ {len(self.fichiers_seq01)} SEQ-01 et {len(self.fichiers_seq02)} SEQ-02 trouvés", foreground=self.colors['success'])
        self.update_status(f"✅ Prêt - {total_fichiers} fichiers trouvés")

    def generer_statistiques(self): # MODIFIÉE pour vérifier répertoire ET sélectionner tout
        if not self.repertoire_parent or not os.path.isdir(self.repertoire_parent):
            messagebox.showwarning("Répertoire Manquant", "Veuillez d'abord sélectionner un répertoire source valide.")
            self.update_status("⚠️ Veuillez sélectionner un répertoire."); return

        selected_items_ids = self.tests_tree.selection()
        if not selected_items_ids:
            self.update_status("ℹ️ Aucun test sélectionné, sélection de tous les tests...")
            self.selectionner_tous_tests()
            self.root.update_idletasks()
            selected_items_ids = self.tests_tree.selection()
            if not selected_items_ids and self.tests_disponibles:
                 messagebox.showerror("Erreur Interne", "Impossible de sélectionner les tests automatiquement."); self.update_status("❌ Erreur sélection auto."); return
            elif not self.tests_disponibles:
                 messagebox.showinfo("Information", "Aucun test disponible."); self.update_status("ℹ️ Aucun test disponible."); return

        self.tests_selectionnes = []
        for item_id in selected_items_ids:
            item_values = self.tests_tree.item(item_id, 'values')
            if not item_values or not isinstance(item_values, (list, tuple)) or len(item_values) == 0: continue 
            identifiant_extraction_test = item_values[0]
            if identifiant_extraction_test.startswith('category_'): continue
            found_test_tuple = next((t for t in self.tests_disponibles if t[2] == identifiant_extraction_test), None)
            if found_test_tuple: self.tests_selectionnes.append(found_test_tuple)
        
        if not self.tests_selectionnes:
            messagebox.showinfo("Information", "Aucun test valide à traiter."); self.update_status("ℹ️ Aucun test valide."); return
        
        self.update_status("🔄 Analyse des fichiers en cours...")
        self.analyser_fichiers()

    def extraire_numero_serie(self, html_content):
        """Identique à l'original"""
        sn_match = re.search(r'Serial Number:</td>\s*<td[^>]*class="hdr_value"[^>]*>\s*([^<]+)\s*</td>', html_content, re.DOTALL)
        if sn_match and sn_match.group(1).strip() != "NONE":
            return sn_match.group(1).strip()
        
        serie_match = re.search(r'série[^<]*</td>\s*<td[^>]*class="value"[^>]*>\s*([^<]+)\s*</td>', html_content, re.DOTALL)
        if serie_match:
            return serie_match.group(1).strip()
        
        return None

    def extraire_statut_test(self, html_content): # Identique
        status_match = re.search(r"UUT Result:.*?<td class='hdr_value'><b><span style=\"color:[^\"]+;\">([^<]+)</span></b></td>", html_content, re.DOTALL | re.IGNORECASE)
        if status_match: return status_match.group(1).strip()
        alt_status_match = re.search(r"UUT Result:.*?<td[^>]*class=\"hdr_value\"[^>]*>.*?<span[^>]*>(Passed|Failed|Terminated)</span>", html_content, re.DOTALL | re.IGNORECASE)
        if alt_status_match: return alt_status_match.group(1).strip()
        return "Inconnu"

    def extraire_date_heure(self, html_content, nom_fichier): # Identique
        date_match = re.search(r"Date:</td>\s*<td[^>]*class=\"hdr_value\"[^>]*><b>([^<]+)</b></td>", html_content, re.DOTALL)
        date_str = date_match.group(1).strip() if date_match else None
        time_match = re.search(r"Time:</td>\s*<td[^>]*class=\"hdr_value\"[^>]*><b>([^<]+)</b></td>", html_content, re.DOTALL)
        heure_str = time_match.group(1).strip() if time_match else None
        if date_str and heure_str:
            try: 
                parts = re.split(r'[\s/]+', date_str) 
                day, month_name, year = parts[-3], parts[-2], parts[-1]
                months_fr_to_num = {"janvier": "01", "février": "02", "mars": "03", "avril": "04", "mai": "05", "juin": "06",
                                     "juillet": "07", "août": "08", "septembre": "09", "octobre": "10", "novembre": "11", "décembre": "12"}
                month_num_str = months_fr_to_num.get(month_name.lower(), month_name)
                date_std = f"{int(day):02d}/{int(month_num_str):02d}/{year}"
                return date_std, heure_str
            except: pass
        return self.extraire_date_heure_du_nom(nom_fichier)

    def extraire_date_heure_du_nom(self, nom_fichier): # Identique
        match = re.search(r'\[(\d{1,2})\s+(\d{1,2})\s+(\d{1,2})\]\[(\d{1,2})\s+(\d{1,2})\s+(\d{4})\]', nom_fichier)
        if match:
            h, m, s, day, month, year = match.groups()
            return f"{int(day):02d}/{int(month):02d}/{year}", f"{int(h):02d}:{int(m):02d}:{int(s):02d}"
        return "date_inconnue", "heure_inconnue"

    def obtenir_cle_tri_chronologique(self, identifiant_unique_avec_date_heure): # Version corrigée pour le tri
        try:
            # Ce regex est crucial. Il s'attend à ce que la partie SN soit TOUT ce qui précède " ["
            match = re.match(r'^(.*?) \[((\d{2})/(\d{2})/(\d{4}))\]\[((\d{2}):(\d{2}):(\d{2}))\]$', identifiant_unique_avec_date_heure)
            if not match:
                sn_part_fallback = identifiant_unique_avec_date_heure.split(' [')[0] if ' [' in identifiant_unique_avec_date_heure else identifiant_unique_avec_date_heure
                return (sn_part_fallback, 0, 0, 0, 0, 0, 0) 

            numero_serie_str_tri = match.group(1).strip() # Ex: "MIQ2022/27-0178" ou "0178"
            
            # Extraire les 4 derniers chiffres pour le tri primaire si possible
            sn_4_chiffres_tri = numero_serie_str_tri
            match_sn_fin = re.search(r'(\d{4})$', numero_serie_str_tri)
            if match_sn_fin:
                sn_4_chiffres_tri = match_sn_fin.group(1)
            elif re.match(r'^\d{1,4}$', numero_serie_str_tri): # Si c'est déjà 1-4 chiffres
                sn_4_chiffres_tri = numero_serie_str_tri.zfill(4)
            # Sinon, sn_4_chiffres_tri reste la chaîne complète pour le tri

            jour_tri = int(match.group(3)) 
            mois_tri = int(match.group(4)) 
            annee_tri = int(match.group(5))
            heure_tri = int(match.group(7))
            minute_tri = int(match.group(8))
            seconde_tri = int(match.group(9))
            
            # Clé de tri: d'abord par les 4 chiffres du SN, puis par la date/heure
            return (sn_4_chiffres_tri, annee_tri, mois_tri, jour_tri, heure_tri, minute_tri, seconde_tri)
        except Exception:
            sn_part_except = identifiant_unique_avec_date_heure.split(' [')[0] if ' [' in identifiant_unique_avec_date_heure else identifiant_unique_avec_date_heure
            return (sn_part_except, 0,0,0,0,0,0)

    def analyser_fichiers(self): # Modifiée pour utiliser le SN brut pour l'ID
        donnees_collectees = {}
        all_files = self.fichiers_seq01 + self.fichiers_seq02
        for idx, fichier_path in enumerate(all_files):
            seq_type = "SEQ-01" if fichier_path in self.fichiers_seq01 else "SEQ-02"
            try:
                with open(fichier_path, "r", encoding="iso-8859-1", errors="replace") as f_html:
                    html_content = f_html.read()
                
                # Utiliser votre fonction extraire_numero_serie originale
                numero_serie_brut_html = self.extraire_numero_serie(html_content) 
                
                # Utiliser le nom du dossier comme fallback si rien n'est trouvé dans le HTML
                if not numero_serie_brut_html:
                    numero_serie_brut_html = os.path.basename(os.path.dirname(fichier_path))
                
                nom_fichier_base = os.path.basename(fichier_path)
                date_test, heure_test = self.extraire_date_heure(html_content, nom_fichier_base)
                statut_test = self.extraire_statut_test(html_content)
                
                # L'identifiant unique utilisera le numero_serie_brut_html
                # C'est cet identifiant qui deviendra l'index et sera affiché en colonne A
                identifiant_unique = f"{numero_serie_brut_html} [{date_test}][{heure_test}]"

                if identifiant_unique not in donnees_collectees:
                    donnees_collectees[identifiant_unique] = {
                        "Numéro de série": numero_serie_brut_html, # Stocker le SN brut pour référence si besoin
                        "Date": date_test, "Heure": heure_test, 
                        "Type": seq_type, "Statut": statut_test
                    }
                else: 
                    donnees_collectees[identifiant_unique]["Type"] += f"/{seq_type}" if seq_type not in donnees_collectees[identifiant_unique]["Type"] else ""

                for nom_complet_test, cat_id, id_extract in self.tests_selectionnes:
                    valeur = None
                    if seq_type == "SEQ-01" and cat_id.startswith("seq01_"):
                        valeur = self.extraire_valeur_seq01(html_content, cat_id, id_extract)
                    elif seq_type == "SEQ-02" and cat_id.startswith("seq02_"):
                        valeur = self.extraire_valeur_seq02(html_content, cat_id, id_extract)
                    if valeur is not None:
                        donnees_collectees[identifiant_unique][nom_complet_test] = valeur
            except Exception as e: print(f"Erreur analyse {fichier_path}: {e}"); traceback.print_exc()
        
        self.data_tests = donnees_collectees
        self.creer_tableau_statistiques()

    def extraire_valeur_seq01(self, html_content, test_parent, identifiant): # Votre version originale
        if test_parent == "seq01_24vdc":
            block_match = re.search(r"Test des alimentations à 24VDC(.*?)(?:Test des alimentations à 115VAC|Calcul des résistances|</body>|$)", html_content, re.DOTALL)
            if not block_match: return None
            block = block_match.group(1)
            if identifiant == "24VDC_+16V": m = re.search(r"Tension \+16V mesurée:\s*</td>\s*<td[^>]*>\s*([-\d\.,]+)", block, re.DOTALL)
            elif identifiant == "24VDC_-16V": m = re.search(r"Tension -16V mesurée:\s*</td>\s*<td[^>]*>\s*([-\d\.,]+)", block, re.DOTALL)
            elif identifiant == "24VDC_+5V": m = re.search(r"Tension \+5V mesurée:\s*</td>\s*<td[^>]*>\s*([-\d\.,]+)", block, re.DOTALL)
            elif identifiant == "24VDC_-5V": m = re.search(r"Tension -5V mesurée:\s*</td>\s*<td[^>]*>\s*([-\d\.,]+)", block, re.DOTALL)
            else: return None
            return m.group(1).replace(',', '.').strip() if m else None
        elif test_parent == "seq01_115vac":
            block_match = re.search(r"Test des alimentations à 115VAC(.*?)(?:Calcul des résistances|</div>|</body>|$)", html_content, re.DOTALL)
            if not block_match: return None
            block = block_match.group(1)
            if identifiant == "115VAC_+16V": m = re.search(r"Tension \+16V mesurée:\s*</td>\s*<td[^>]*>\s*([-\d\.,]+)", block, re.DOTALL)
            elif identifiant == "115VAC_-16V": m = re.search(r"Tension -16V mesurée:\s*</td>\s*<td[^>]*>\s*([-\d\.,]+)", block, re.DOTALL)
            else: return None
            return m.group(1).replace(',', '.').strip() if m else None
        elif test_parent == "seq01_resistances":
            if identifiant == "R46_calculee": r_match = re.search(r"Résistance R46 calculée:\s*</td>\s*<td[^>]*>\s*([\d\.,]+)\s*</td>", html_content, re.DOTALL)
            elif identifiant == "R46_monter": r_match = re.search(r"Résistance R46 à monter:\s*</td>\s*<td[^>]*>\s*Résistance à monter =\s*([\d]+)\s*ohms", html_content, re.DOTALL); return r_match.group(1).strip() if r_match else None
            elif identifiant == "R47_calculee": r_match = re.search(r"Résistance R47 calculée:\s*</td>\s*<td[^>]*>\s*([\d\.,]+)\s*</td>", html_content, re.DOTALL)
            elif identifiant == "R47_monter": r_match = re.search(r"Résistance R47 à monter:\s*</td>\s*<td[^>]*>\s*Résistance à monter =\s*([\d]+)\s*ohms", html_content, re.DOTALL); return r_match.group(1).strip() if r_match else None
            elif identifiant == "R48_calculee": r_match = re.search(r"Résistance R48 calculée:\s*</td>\s*<td[^>]*>\s*([\d\.,]+)\s*</td>", html_content, re.DOTALL)
            elif identifiant == "R48_monter": r_match = re.search(r"Résistance R48 à monter:\s*</td>\s*<td[^>]*>\s*Résistance à monter =\s*([\d]+)\s*ohms", html_content, re.DOTALL); return r_match.group(1).strip() if r_match else None
            else: return None
            return str(int(round(float(r_match.group(1).replace(',', '.'))))) if r_match else None # Pour RXX_calculee
        return None

    def extraire_valeur_seq02(self, html_content, test_parent, identifiant): # Votre version originale
        if test_parent == "seq02_transfert":
            if identifiant == "Test_19VDC":
                block_match = re.search(r"Test 1\.9Un sur 2 voies en 19VDC.*?(Mesure -16V en V:.*?</table>)", html_content, re.DOTALL)
                if block_match: m = re.search(r"Mesure -16V en V:\s*</td>\s*<td[^>]*>\s*([-\d\.,]+)", block_match.group(1), re.DOTALL); return m.group(1).replace(',', '.').strip() if m else None
            elif identifiant == "Test_115VAC":
                block_match = re.search(r"Test 1\.9Un sur 2 voies en 115VAC.*?(Mesure -16V en V:.*?</table>)", html_content, re.DOTALL)
                if block_match: m = re.search(r"Mesure -16V en V:\s*</td>\s*<td[^>]*>\s*([-\d\.,]+)", block_match.group(1), re.DOTALL); return m.group(1).replace(',', '.').strip() if m else None
        elif test_parent == "seq02_precision_transfert" and identifiant == "precision_transfert_defauts":
            defauts = extraire_defauts_precision_transfert(html_content)
            return "; ".join([f"{nom}: {precision}" for nom, precision in defauts]) if defauts else ""
        return None

    def creer_tableau_statistiques(self): # MODIFIÉE pour le tri et XLSM
        if not self.data_tests: messagebox.showinfo("Information", "Aucune donnée collectée."); self.update_status("ℹ️ Aucune donnée."); return
        df_full = pd.DataFrame.from_dict(self.data_tests, orient='index')
        if df_full.empty: messagebox.showinfo("Information", "Aucune donnée à afficher."); self.update_status("ℹ️ DataFrame vide."); return
            
        if self.tri_chrono_var.get():
            serie_de_tri = pd.Series([self.obtenir_cle_tri_chronologique(idx) for idx in df_full.index], index=df_full.index)
            df_full = df_full.loc[serie_de_tri.sort_values().index]
        else:
            df_full = df_full.sort_index()

        cols_to_drop = ["Numéro de série", "Date", "Heure"] # Colonnes originales de self.data_tests
        df_to_write = df_full.drop(columns=[col for col in cols_to_drop if col in df_full.columns], errors='ignore')

        ordered_cols = ["Type", "Statut"]
        if "Défauts précision transfert" in df_to_write.columns: ordered_cols.append("Défauts précision transfert")
        selected_test_names = [test[0] for test in self.tests_selectionnes if test[0] in df_to_write.columns]
        for col_name in selected_test_names:
            if col_name not in ordered_cols: ordered_cols.append(col_name)
        remaining_cols = [col for col in df_to_write.columns if col not in ordered_cols]
        final_ordered_cols = ordered_cols + remaining_cols
        df_to_write = df_to_write[final_ordered_cols]
        
        # --- DIAGNOSTIC ---
        print("--- DIAGNOSTIC creer_tableau_statistiques ---")
        print("Index de df_to_write (pour colonne A Excel):")
        print(df_to_write.index)
        print("Premières lignes de df_to_write:")
        print(df_to_write.head().to_string())
        print("--- FIN DIAGNOSTIC ---")

        nom_fichier_xlsm = "statistiques_SEQ01_SEQ02.xlsm"
        chemin_xlsm = os.path.join(self.repertoire_parent, nom_fichier_xlsm)
        template_path = resource_path("template_statistiques.xlsm")

        try:
            self.update_status("📊 Génération XLSM...")
            wb = openpyxl.load_workbook(template_path, keep_vba=True)
            ws_data = wb["Statistiques_SEQ01_02"] if "Statistiques_SEQ01_02" in wb.sheetnames else wb.create_sheet("Statistiques_SEQ01_02", 0)
            if ws_data.max_row > 1:
                for r in range(ws_data.max_row, 1, -1): ws_data.delete_rows(r)
            
            header_excel = [df_to_write.index.name or "Identifiant Test"] + list(df_to_write.columns)
            ws_data.append(header_excel)
            for cell_h in ws_data[1]: cell_h.font = Font(bold=True)
            
            for idx_val, data_r in df_to_write.iterrows():
                ws_data.append([str(idx_val)] + list(data_r.values))

            ws_config = wb["ConfigSheet"] if "ConfigSheet" in wb.sheetnames else wb.create_sheet("ConfigSheet")
            ws_config['A1'] = self.repertoire_parent
            
            ws_data.auto_filter.ref = ws_data.dimensions
            for col_obj in ws_data.columns:
                max_l = 0; letter = col_obj[0].column_letter
                for cell_obj_dim in col_obj:
                    if cell_obj_dim.value: max_l = max(max_l, len(str(cell_obj_dim.value)))
                ws_data.column_dimensions[letter].width = max_l + 5 if max_l > 0 else 15

            map_cols_format = {c.value: c.column for c in ws_data[1] if c.value}
            for r_num in range(2, ws_data.max_row + 1):
                for nom_c, (b_min, b_max) in bornes_alims_seq01.items():
                    c_idx = map_cols_format.get(nom_c)
                    if c_idx:
                        cell_f = ws_data.cell(row=r_num, column=c_idx)
                        try:
                            if cell_f.value is not None:
                                val_s = str(cell_f.value).replace(",",".").strip()
                                if val_s:
                                    val_f = float(val_s)
                                    if not (b_min <= val_f <= b_max): cell_f.font = Font(color="FF0000")
                        except (ValueError,TypeError): pass
                for pfx in ["R46","R47","R48"]:
                    c_calc_k, c_mont_k = f"{pfx} calculée", f"{pfx} à monter"
                    c_calc_idx, c_mont_idx = map_cols_format.get(c_calc_k), map_cols_format.get(c_mont_k)
                    if c_calc_idx and c_mont_idx:
                        v_calc, v_mont = ws_data.cell(r_num,c_calc_idx).value, ws_data.cell(r_num,c_mont_idx).value
                        try:
                            if v_calc is not None and v_mont is not None:
                                if int(float(str(v_mont).replace(",","."))) > int(float(str(v_calc).replace(",","."))):
                                    ws_data.cell(r_num,c_mont_idx).font = Font(color="FF0000")
                        except (ValueError,TypeError): pass
            
            if ws_data.max_row > 1 and ws_data.max_column > 1: ws_data.freeze_panes = ws_data['B2']
            elif ws_data.max_row > 1: ws_data.freeze_panes = ws_data['A2']
            elif ws_data.max_column > 1: ws_data.freeze_panes = ws_data['B1']

            wb.save(chemin_xlsm)
            self.update_status(f"✅ Rapport XLSM: {os.path.basename(chemin_xlsm)}")
            if sys.platform == "win32": os.startfile(chemin_xlsm)
            elif sys.platform == "darwin": subprocess.call(["open", chemin_xlsm])
            else: subprocess.call(["xdg-open", chemin_xlsm])
        except Exception as e_final: messagebox.showerror("Erreur Finale XLSM", f"{e_final}"); traceback.print_exc()

    def generer_rapports_detailles(self): # Identique
        if not self.repertoire_parent: messagebox.showinfo("Information", "Veuillez sélectionner un répertoire."); return
        try:
            self.update_status("🔍 Recherche sous-répertoires...")
            sous_repertoires = [os.path.join(self.repertoire_parent, d) for d in os.listdir(self.repertoire_parent) if os.path.isdir(os.path.join(self.repertoire_parent, d))]
            if not sous_repertoires: messagebox.showinfo("Information", "Aucun sous-répertoire trouvé."); self.update_status("ℹ️ Aucun sous-répertoire."); return
            self.update_status(f"📄 Génération rapports pour {len(sous_repertoires)} séries...")
            progress_window = ProgressWindow(total_files=len(sous_repertoires))
            for repertoire in sous_repertoires: traiter_repertoire_serie(repertoire, progress_window)
            progress_window.show_completion(); self.root.wait_window(progress_window.window)
            self.update_status("✅ Rapports détaillés générés.")
        except Exception as e: messagebox.showerror("Erreur", f"Erreur rapports détaillés: {e}"); traceback.print_exc()

    def lancer(self): self.root.mainloop()

if __name__ == "__main__":
    app = ModernStatsTestsWindow()
    app.lancer()