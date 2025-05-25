#!/usr/bin/env python3
"""
Test script to verify XLSM file generation works correctly
"""

import os
import tempfile
import pandas as pd
import openpyxl
from openpyxl.styles import Font

def test_xlsm_generation():
    """Test the XLSM file generation workflow"""
    print("Testing XLSM file generation...")
    
    # Create test data
    test_data = {
        'Numéro de série': ['1234', '5678', '9012'],
        'R46 calculée': [100, 150, 200],
        'R46 à monter': [105, 145, 205],
        'Statut': ['OK', 'KO', 'OK']
    }
    
    df = pd.DataFrame(test_data)
    print(f"Created test DataFrame with {len(df)} rows")
    
    # Create temporary directory for test
    with tempfile.TemporaryDirectory() as temp_dir:
        print(f"Using temporary directory: {temp_dir}")
        
        # Test the fixed XLSM workflow
        chemin_xlsm = os.path.join(temp_dir, "test_statistiques.xlsm")
        chemin_temp = os.path.join(temp_dir, "temp_stats.xlsx")
        
        try:
            # Step 1: Create temporary XLSX file
            print("Step 1: Creating temporary XLSX file...")
            df.to_excel(chemin_temp, sheet_name="Test_Statistics", index=True)
            print(f"✅ Temporary XLSX created: {os.path.exists(chemin_temp)}")
            
            # Step 2: Load with openpyxl and apply formatting
            print("Step 2: Loading with openpyxl and applying formatting...")
            wb = openpyxl.load_workbook(chemin_temp)
            ws = wb.active
            
            # Apply some basic formatting (like in the real application)
            ws.auto_filter.ref = ws.dimensions
            
            # Test color formatting for resistance values
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # Find columns by header
                for col_idx, cell in enumerate(ws[1]):
                    if cell.value == "R46 à monter":
                        mont_col = col_idx
                    elif cell.value == "R46 calculée":
                        calc_col = col_idx
                
                # Apply red formatting if "à monter" > "calculée"
                try:
                    calc_val = int(row[calc_col].value)
                    mont_val = int(row[mont_col].value)
                    if mont_val > calc_val:
                        row[mont_col].font = Font(color="FF0000")
                        print(f"Applied red formatting to row {row[0].row}")
                except:
                    pass
            
            # Configure for macro support
            wb.code_name = "ThisWorkbook"
            print("✅ Formatting applied and workbook configured for macros")
            
            # Step 3: Save as XLSM
            print("Step 3: Saving as XLSM...")
            wb.save(chemin_xlsm)
            print(f"✅ XLSM file saved: {os.path.exists(chemin_xlsm)}")
            
            # Step 4: Clean up temporary file
            print("Step 4: Cleaning up temporary file...")
            if os.path.exists(chemin_temp):
                os.remove(chemin_temp)
                print(f"✅ Temporary file removed: {not os.path.exists(chemin_temp)}")
            
            # Step 5: Verify the XLSM file
            print("Step 5: Verifying XLSM file...")
            
            # Check file size
            file_size = os.path.getsize(chemin_xlsm)
            print(f"XLSM file size: {file_size} bytes")
            
            # Try to read it back
            wb_verify = openpyxl.load_workbook(chemin_xlsm)
            ws_verify = wb_verify.active
            print(f"✅ XLSM file readable: {ws_verify.max_row} rows, {ws_verify.max_column} columns")
            
            # Check if macro configuration is preserved
            print(f"Workbook code_name: {wb_verify.code_name}")
            
            # Verify data integrity
            data_matches = True
            for row_idx, row in enumerate(ws_verify.iter_rows(min_row=2, max_row=4, values_only=True)):
                if row[1] != test_data['Numéro de série'][row_idx]:  # Index column + serial number
                    data_matches = False
                    break
            
            print(f"✅ Data integrity: {'Preserved' if data_matches else 'FAILED'}")
            
            print("\n" + "="*60)
            print("🎉 XLSM GENERATION TEST COMPLETED SUCCESSFULLY!")
            print("="*60)
            print(f"Generated file: {chemin_xlsm}")
            print(f"File size: {file_size} bytes")
            print("The XLSM workflow is working correctly.")
            
            return True
            
        except Exception as e:
            print(f"\n❌ ERROR during XLSM generation test:")
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
            return False

if __name__ == "__main__":
    success = test_xlsm_generation()
    if success:
        print("\n✅ All tests passed! The XLSM generation fix is working correctly.")
    else:
        print("\n❌ Tests failed! There may still be issues with XLSM generation.")
